import openpyxl
import re
import time
import helper
import run_process
import shutil
import os

def parsing(stdFilePath):
	#Get directory
	#pathdir = os.path.dirname(os.path.abspath(__file__))
	pathdir = "D:\PYTHON\STDF_TO_ASCII_CONVERTER"
	print(pathdir)
	
	#Get std file name
	stdname = re.search(r"\w+.std", stdFilePath).group(0)
	print(stdname)
	
	#Copy std file to folder
	targetpath = pathdir + "\\" + stdname
	print(targetpath)
	shutil.copyfile(stdFilePath, targetpath)
	
	#Convert STD file to text file
	run_process.OpenSTDFReader(stdname, pathdir)
	
	#Configure time
	timestamp = time.ctime()
	timestamp = timestamp.replace("Wed","")
	timestamp = timestamp.replace(" ","_")
	timestamp = timestamp.replace(":","_")
	
	#Create workbook
	workbookPath = stdFilePath.strip(".std") + "_Parsed_" + timestamp + ".xlsx"
	openpyxl.Workbook().save(workbookPath)
	workbookActive = openpyxl.load_workbook(workbookPath)
	
	#Changing the name of the sheet
	sheetlist = workbookActive.sheetnames
	sheetname = sheetlist[0]
	sheet = workbookActive[sheetname]
	sheet.title = "PARSED OUTPUT"
	workbookActive.create_sheet("MIR INFO", 0)
	sheetMIR = workbookActive["MIR INFO"]
	
	#Text path
	txt_path = targetpath.strip(".std") + (".txt")
	print(txt_path)
	
	#Open text file
	stdffile = open(txt_path, 'r')
	
	#Get contents of the text file
	contentList = []
	stdffile.seek(0)
	for content in stdffile.readlines():
		contentList.append(content)
	maxLine = len(contentList)
	
	#Convert content to string so we can find searches
	contentString = helper.ListToString(contentList)
	
	#Number of attributes per Record
	rangeMIR = 41
	rangePTR = 24
	rangePRR = 15
	rangePIR = 5
	spacePTR_PRR = 2
	spacePRR_PIR = 3
	spacePIR_TR = 3

	#Get indexes of the PTR matches
	ptrStr = "                Parametric Test Result Record\n"
	ptrList = helper.FindIndexDuplicate(contentList, ptrStr)
	ptrList = helper.RemoveDuplicate(ptrList)

	#Get indexes of the PRR matches
	prrStr = "                Part Result Record\n"
	prrList = helper.FindIndexDuplicate(contentList, prrStr)
	prrList = helper.RemoveDuplicate(prrList)
	
	#Get MIR index
	mirMatch = re.findall("                Master Information Record\n", contentString)
	mirIndex = contentList.index(mirMatch[0]) + 1
	
	#Containers
	stringvar = ''
	rowlist = []
	dummylist = []
	
	#Convert MIR group string into list
	MIRlist = []
	MIRInfo = []
	for row in contentList[mirIndex:mirIndex +rangeMIR]:
		for elem in row:
			stringvar += elem
		dummylist = stringvar.strip('\n').split(' ')
		for elem in dummylist:
			if elem != '':
				rowlist.append(elem)
		#Append list to MIRlist
		MIRlist.append(rowlist)
		#Empty lists and string
		dummylist = []
		rowlist = []
		stringvar = ''
	for row in MIRlist:
		if (len(row)>2) and (row[2] == "not"):
			row[2] = ''
			del row[3]
		del row[1]
	MIRlist[11][0] = "LOT ID"
	MIRInfo.append(MIRlist[11]) # LOT ID
	MIRlist[12][0] = "PART TYPE"
	MIRInfo.append(MIRlist[12]) # PART TYPE
	MIRlist[25][0] = "PACKAGE TYPE"
	MIRInfo.append(MIRlist[25]) # PACKAGE TYPE
	MIRlist[32][0] = "SPEC TYPE"
	MIRInfo.append(MIRlist[32]) # DEVICE TYPE
	MIRlist[33][0] = "SPEC REVISION"
	MIRInfo.append(MIRlist[33]) # SPEC VERSION
	MIRlist[15][0] = "TEST JOB"
	MIRInfo.append(MIRlist[15]) # TEST JOB
	MIRlist[16][0] = "TEST JOB REVISION"
	MIRInfo.append(MIRlist[16]) # TEST JOB REVISION
	MIRlist[34][0] = "TEST FLOW ID"
	MIRInfo.append(MIRlist[34]) # FLOW ID
	MIRlist[14][0] = "TESTER"
	MIRInfo.append(MIRlist[14]) # TESTER
	MIRlist[20][0] = "TESTXENTRAL VERSION"
	MIRInfo.append(MIRlist[20]) # TESTXENTRAL VERSION
	
	mirname = ["MASTER INFORMATION RECORD"]
	MIRInfo.insert(0, mirname)
	#Copy info to excel
	for row in MIRInfo:
		sheetMIR.append(row)
	sheetMIR.merge_cells('A1:B1')
	
	#Get PTR VALUES
	PRRvalues = []
	PTRlist = []
	PTRdict = {}
	TESTdict = {}
	testnamelist = []
	loopflag = 0
	prrLoop = 0
	ptrLoop = 0
	sitenum = 0
	for loop in range(len(ptrList)): #len(ptrList)
		#Convert PTR group string into list
		for row in contentList[ptrList[loop]+1:ptrList[loop]+rangePTR]:
			for elem in row:
				stringvar += elem
			dummylist = stringvar.strip('\n').split(' ')
			for elem in dummylist:
				if elem != '':
					rowlist.append(elem)
			#Append list to PTRlist
			PTRlist.append(rowlist)
			#Empty lists and string
			dummylist = []
			rowlist = []
			stringvar = ''
		
		#Keep values into dictionary
		if PTRlist[5][2] == '255':
			testnamelist.append(PTRlist[3][2] + '_' + PTRlist[9][2] + '_' + PTRlist[9][3])
			TESTdict[PTRlist[3][2] + '_' + PTRlist[9][2] + '_' + PTRlist[9][3]] = ''
		else:
			if PTRlist[3][2] + '_' + PTRlist[9][2] + '_' + PTRlist[9][3] in TESTdict:
				TESTdict[PTRlist[3][2] + '_' + PTRlist[9][2] + '_' + PTRlist[9][3]] = PTRlist[8][2]
				#TESTdict[PTRlist[3][2] + '_' + PTRlist[9][2] + '_' + PTRlist[9][3]] = helper.ChangeUnit(PTRlist[8][2],PTRlist[12][2])
				sitenum = PTRlist[5][2]
		
		if ((ptrList[ptrLoop] + (rangePTR) + spacePTR_PRR) == prrList[prrLoop]):
			#Get PRR values
			for row in contentList[prrList[prrLoop]+1:prrList[prrLoop] +rangePRR]:
				for elem in row:
					stringvar += elem
				dummylist = stringvar.strip('\n').split(' ')
				for elem in dummylist:
					if elem != '':
						rowlist.append(elem)
				#Append list to PRRlist
				PRRvalues.append(rowlist)
				#Empty lists and string
				dummylist = []
				rowlist = []
				stringvar = ''
			
			dummy2list = []
			keysTest = TESTdict.keys()
			for keys in keysTest:
				dummy2list.append(TESTdict[keys])
			dummy2list.insert(0, prrLoop)	#touchdown
			dummy2list.insert(1, sitenum)	#sitenum
			dummy2list.insert(2, PRRvalues[9][2])	#diex
			dummy2list.insert(3, PRRvalues[10][2])	#diey
			dummy2list.insert(4, PRRvalues[8][2])	#soft bin
			dummy2list.insert(5, PRRvalues[7][2])	#hard bin
			PTRdict[prrLoop] = dummy2list
			
			#Empty variables
			keyslist = TESTdict.keys()
			for keys in keyslist:
				TESTdict[keys] = ''
			#Increment PRR index
			prrLoop += 1
			
		#Empty PTRlist
		PTRlist = []
			
		#Increment PTR index 
		ptrLoop += 1
	
	
	#Sort testnames
	testnamelist = helper.RemoveDuplicate(testnamelist)
	testnamelist.insert(0, "Loop Num")
	testnamelist.insert(1, "Site Num")
	testnamelist.insert(2, "DIE X")
	testnamelist.insert(3, "DIE Y")
	testnamelist.insert(4, "Soft Bin")
	testnamelist.insert(5, "Hard Bin")
	sheet.append(testnamelist)
	
	#Copy lists to excel
	keyslist = PTRdict.keys()
	for keys in keyslist:
		if keys != 0:
			sheet.append(PTRdict[keys])
	
	#Save excel file
	workbookActive.save(workbookPath)
	
	#Delete files
	stdffile.close()
	os.remove(targetpath)
	os.remove(txt_path)